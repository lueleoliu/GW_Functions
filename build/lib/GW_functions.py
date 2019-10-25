# -*- coding: utf-8 -*-
# Bladed 4.6 #
# In文件替换与基本参数输出 #
# 风文件空气密度提取，最小桨距角评估与Kopt计算 #
# 针对模型线性化及控制参数修正 #

import xlrd
import os
import shutil
import math
import time
import csv
from openpyxl import Workbook
import win32com.client
from lxml import etree
from multiprocessing.dummy import Pool as ThreadPool
import tkinter
import tkinter.messagebox
import pandas


# Function 空气密度信息提取 # Done
def get_airdensity(root_dir):

    wind_dir = os.path.join(root_dir, 'WindFile')
    windfile_dir = get_typefile(wind_dir, '.xlsx')
    isxlsx = 1

    if not windfile_dir:
        windfile_dir = get_typefile(wind_dir, '.txt')
        isxlsx = 0

    if len(windfile_dir) > 1:
        logging(root_dir, ' ERROR : 风文件过多！')
        tkinter.messagebox.showerror('错误', 'ERROR 101: 风文件过多！')
        raise Exception('风文件过多！')

    elif not windfile_dir:
        logging(root_dir, ' ERROR : 没有风文件数据！')
        tkinter.messagebox.showerror('错误', 'ERROR 102: 没有风文件数据！')
        raise Exception('没有风文件数据')

    if isxlsx:
        wind = xlrd.open_workbook(windfile_dir[0])
        site_condition = wind.sheet_by_name('Site Condition')
        airdensity = site_condition.col_values(6)
        airdensity = [i for i in airdensity if isinstance(i, float)]

    else:
        with open(windfile_dir[0], 'r') as wind:
            lines = wind.readlines()
            airdensity = float(lines[0])

    return max(airdensity)


# Function 基础in文件生成 # Done
# 默认按照Cp计算进行in文件生成 #
# 修正空气密度 #
# 修改控制器路径 #
def gen_originalin(root_dir):
    global logfile

    model_dir = os.path.join(root_dir, 'Model')
    project_dir = get_typefile(model_dir, '.prj')

    if not project_dir:
        logging(root_dir, ' ERROR : 未找到模型文件！')
        tkinter.messagebox.showerror('错误', 'ERROR 103: 未找到模型文件！')
        raise Exception('未找到模型文件')

    control_dir = os.path.join(root_dir, 'Exctrl')
    dll_file_path = get_typefile(control_dir, '.dll')
    xml_file_path = get_typefile(control_dir, '.xml')

    change_xml(project_dir[0], 'ExternalController', 'Filepath', dll_file_path[0])
    change_xml(project_dir[0], 'ExternalController', 'AdditionalParameters', 'READ ' + xml_file_path[0])

    change_info(project_dir[0], 'CALCULATION', '2')
    run_bat(root_dir, 'Model')
    in_dir = get_typefile(model_dir, '.in')[0]
    eigen_b = catch_block(in_dir, 'EIGENB')
    eigen_b.append('\n')
    eigen_t = catch_block(in_dir, 'EIGENT')
    eigen_t.append('\n')

    delete_block(project_dir[0], 'RMODE')
    delete_info(project_dir[0], '0RMASS')

    change_info(project_dir[0], 'CALCULATION', '10')
    change_block(project_dir[0], 'CONSTANTS', 'RHO', str(get_airdensity(root_dir)))
    run_bat(root_dir, 'Model')

    add_block(in_dir, 'RCON', eigen_b)
    add_block(in_dir, 'RCON', eigen_t)

    logging(root_dir, ' 初始in文件生成成功')
    print('初始in文件生成成功')

    return


# subFunction Cp计算in文件生成并执行计算 # Done
# 修正桨距角为-2至2度，步长0.5度 #
def gen_performance(root_dir):
    global logfile
    gen_originalin(root_dir)

    ori_in_dir = os.path.join(root_dir, 'Model')
    ori_in_file_path = get_typefile(ori_in_dir, '.in')
    current_in_dir = os.path.join(root_dir, 'Performance')
    current_in_file_path = ori_in_file_path[0].replace(ori_in_dir, current_in_dir)
    shutil.copyfile(ori_in_file_path[0], current_in_file_path)

    change_info(current_in_file_path, 'CALCN', '5')
    change_info(current_in_file_path, 'PATH', current_in_dir)
    change_info(current_in_file_path, 'RUNNAME', 'pcoeffs')
    change_info(current_in_file_path, 'OPTNS', '0')
    change_block(current_in_file_path, 'PCOEFF', 'PITCH', '-0.03490660')
    change_block(current_in_file_path, 'PCOEFF', 'PITCH_END', '0.03490660')
    change_block(current_in_file_path, 'PCOEFF', 'PITCH_STEP', '0.00872665')

    print('开始计算最小桨距角')
    run_bat(root_dir, 'Performance')
    logging(root_dir, ' 最小桨距角计算完成')
    print('最小桨距角计算完成')

    return


# subFunction 提取最优Cp及对应的叶尖速比与最小桨距角 # Done
def get_cpinfo(root_dir):
    cpinfo = []
    cp_path = os.path.join(root_dir, 'Performance')
    data_path = get_typefile(cp_path, '.%37')
    with open(data_path[0], 'r') as msg:
        lines = msg.readlines()
        for j in range(len(lines)):
            if 'ULOADS' in lines[j]:
                cpinfo.append(do_split(lines[j], '  ', 1))     # 提取Cp
                cpinfo.append(do_split(lines[j], '  ', 2))     # 提取λ
            elif 'MAXTIME'in lines[j]:
                cpinfo.append(do_split(lines[j], '  ', 1))     # 提取最小桨距角
    return cpinfo


# subFunction Kopt计算 # Done
def get_optmodegain(root_dir):
    global logfile

    cpinfo = get_cpinfo(root_dir)
    cp_max = float(cpinfo[0].strip())
    lamda = float(cpinfo[1].strip())

    model_dir = os.path.join(root_dir, 'Model')
    project_dir = get_typefile(model_dir, '.prj')
    radius = 0.5 * get_block(project_dir[0], 'RCON', 'DIAM')

    pho = get_block(project_dir[0], 'CONSTANTS', 'RHO')

    k_opt = math.pi * pho * math.pow(radius, 5) * cp_max / (2 * math.pow(lamda, 3))
    logging(root_dir, ' Kopt计算完成')
    print('Kopt计算完成')
    return int(k_opt)


# Function 模板In文件生成 # Done
# 修改最小桨距角及Kopt #
# 不执行计算 #
def gen_standard(root_dir):
    global logfile
    gen_performance(root_dir)

    ori_in_dir = os.path.join(root_dir, 'Model')
    ori_in_file_path = get_typefile(ori_in_dir, '.in')

    change_block(ori_in_file_path[0], 'CONTROL', 'GAIN_TSR', str(get_optmodegain(root_dir)))
    change_block(ori_in_file_path[0], 'CONTROL', 'PITMIN', get_cpinfo(root_dir)[2])

    logging(root_dir, ' 修正in文件完成')
    print('修正in文件完成')

    return


# Function Campbell图in文件生成并计算 # Done
# 前置计算：gen_standard #
def gen_campbell(root_dir):
    global logfile

    ori_in_dir = os.path.join(root_dir, 'Model')
    ori_in_file_path = get_typefile(ori_in_dir, '.in')
    current_in_dir = os.path.join(root_dir, 'Campbell')
    current_in_file_path = ori_in_file_path[0].replace(ori_in_dir, current_in_dir)
    shutil.copyfile(ori_in_file_path[0], current_in_file_path)

    change_info(current_in_file_path, 'CALCN', '17')
    change_info(current_in_file_path, 'PATH', current_in_dir)
    change_info(current_in_file_path, 'RUNNAME', 'Campbell')
    change_info(current_in_file_path, 'OPTNS', '144658')

    cut_in = get_block(current_in_file_path, 'RCON', 'CUTIN')
    cut_out = get_block(current_in_file_path, 'RCON', 'CUTOUT')

    block_lin = ['\n',
                 'MSTART LINEARISE\n',
                 'LINTYPE\t1\n',
                 'WINDLO\t' + str(cut_in) + '\n',
                 'WINDHI\t' + str(cut_out) + '\n',
                 'WINDSTEP\t.5\n',
                 'AZLO\t0\n', 'AZHI\t0\n',
                 'AZSTEP\t1.745328E-02\n',
                 'MAXFREQ\t1\n',
                 'CCCRIT\t.8\n',
                 'LINPATH\t..\n',
                 'MBCFLAG\t1\n',
                 'MEND\n',
                 '\n']

    add_block(current_in_file_path, 'AEROINFO', block_lin)

    change_xml(current_in_file_path, 'PitchActuator', 'DemandType', 'Position')
    change_xml(current_in_file_path, 'PitchActuator', 'HasLimitSwitches', 'false')
    change_xml(current_in_file_path, 'PitchActuator', 'IncludeBearingFriction', 'false')
    delete_xml(current_in_file_path, 'PitchActuator', 'Drive', 'type', 'ActuatorDrive_None')

    print('开始计算Campbell')
    run_bat(root_dir, 'Campbell')
    logging(root_dir, ' Campbell 文件生成完毕')
    print('Campbell 文件生成完毕')

    return


# Function 线性化in文件生成并计算 # Done
def gen_linear(root_dir):
    global logfile

    ori_in_dir = os.path.join(root_dir, 'Campbell')
    ori_in_file_path = get_typefile(ori_in_dir, '.in')
    current_in_dir = os.path.join(root_dir, 'Linear')
    current_in_file_path = ori_in_file_path[0].replace(ori_in_dir, current_in_dir)
    shutil.copyfile(ori_in_file_path[0], current_in_file_path)

    change_info(current_in_file_path, 'CALCN', '16')
    change_info(current_in_file_path, 'PATH', current_in_dir)
    change_info(current_in_file_path, 'RUNNAME', 'lin1')

    delete_block(current_in_file_path, 'LINEARISE')

    cut_in = get_block(current_in_file_path, 'RCON', 'CUTIN')
    cut_out = get_block(current_in_file_path, 'RCON', 'CUTOUT')

    block_lin = ['\n',
                 'MSTART LINEARISE\n',
                 'LINTYPE\t0\n',
                 'WINDLO\t' + str(cut_in) + '\n',
                 'WINDHI\t' + str(cut_out) + '\n',
                 'WINDSTEP\t.5\n',
                 'AZLO\t0\n', 'AZHI\t0\n',
                 'AZSTEP\t1.745328E-02\n',
                 'WINDPERT\t.1\n',
                 'INDIVWIND\t0\n',
                 'PITCHPERT\t8.726639E-03\n',
                 'INDIVPITCH\t0\n',
                 'QGPERT\t30\n',
                 'MBCFLAG\t1\n',
                 'MEND\n',
                 '\n']

    add_block(current_in_file_path, 'AEROINFO', block_lin)
    print('开始执行线性化计算')
    run_bat(root_dir, 'Linear')
    logging(root_dir, ' 线性化计算完成')
    print('线性化计算完成')

    return


# Function 线性化模型导出in文件生成并计算 #
# 前置计算：gen_campbell #
def gen_linear_model(root_dir):
    global logfile
    gen_linear(root_dir)

    lin_path = os.path.join(root_dir, 'Linear')
    linmod_path = os.path.join(root_dir, 'LinearModel')

    var_abbrev = []
    var_name = []

    var_abbrev.append('Blade 1 pitch angle')
    var_name.append('Blade 1 pitch angle')

    var_abbrev.append('Blade 1 pitch rate')
    var_name.append('Blade 1 pitch rate')

    var_abbrev.append('Nacelle fore-aft acceleration')
    var_name.append('Nacelle fore-aft acceleration')

    var_abbrev.append('Nacelle side-side acceleration')
    var_name.append('Nacelle side-side acceleration')

    var_abbrev.append('Generator torque')
    var_name.append('Generator torque')

    var_abbrev.append('Generator speed')
    var_name.append('Generator speed')

    var_abbrev.append('Measured generator speed')
    var_name.append('Measured generator speed')

    var_abbrev.append('Blade 1 demanded pitch angle')
    var_name.append('Blade 1 demanded pitch angle')

    # 扩展：可以遍历所有数据索引文件（*.%xx），取出变量名与所在文件后缀（数字），组成该dict #
    var_exten = {'Measured generator speed': '04',
                 'Generator speed': '05',
                 'Generator torque': '06',
                 'Blade 1 pitch angle': '08',
                 'Blade 1 pitch rate': '08',
                 'Blade 1 demanded pitch angle': '08',
                 'Nacelle fore-aft acceleration': '26',
                 'Nacelle side-side acceleration': '26'}

    linmod_do = ['SDSTAT\t2\n',
                 'PATH\t' + linmod_path + '\n',
                 'RUNNAME\tlinmod1\n',
                 'MSTART LINMOD\n'
                 'INPATH\t' + lin_path + '\n',
                 'INNAME\tlin1\n',
                 'NOUTPUTS\t' + str(len(var_name)) + '\n',
                 'CCCRIT\t.8\n',
                 'CCDISP 0 0\n',
                 'OUTPUTSTYLE\t0\n',
                 'MEND\n',
                 '\n']

    var_do = []

    for i in range(len(var_name)):
        var_do.append("EXTEN\t" + var_exten[var_name[i]] + "\n")
        var_do.append("ABBREV\t'" + var_abbrev[i] + "'\n")
        var_do.append("NAME\t'" + var_name[i] + "'\n")

    linmod_do[6:6] = var_do

    with open(os.path.join(linmod_path, 'DTLINMOD.IN'), 'w+') as linmod:
        linmod.writelines(linmod_do)

    run_bat(root_dir, 'LinearModel')

    logging(root_dir, ' 线性化结果提取完成')
    print('线性化结果提取完成')

    return


# Function 风机基本信息提取至控制器 #
def get_wt_basic_info(root_dir):

    in_dir = os.path.join(root_dir, 'Model')
    in_file_path = get_typefile(in_dir, '.in')[0]

    control_dir = os.path.join(root_dir, 'Exctrl')
    xml_file_path = get_typefile(control_dir, '.xml')[0]

    rated_speed = get_block(in_file_path, 'CONTROL', 'OMDEM_QS')
    rated_torque = get_block(in_file_path, 'CONTROL', 'GTORREF')
    rated_shaft_power = rated_speed * rated_torque
    min_speed = get_block(in_file_path, 'CONTROL', 'OMMIN')

    controller_edit(xml_file_path, 'Channel', 'P_FaultPitchRunawayRate',
                    str(get_block(in_file_path, 'CONTROL', 'PITRMIN')))

    controller_edit(xml_file_path, 'Channel', 'P_RatedGeneratorSpeed',
                    str(rated_speed))
    controller_edit(xml_file_path, 'Channel', 'P_SteadyShaftPowerLimit',
                    str(rated_shaft_power))
    controller_edit(xml_file_path, 'Channel', 'P_MinimumFinePitch',
                    str(get_block(in_file_path, 'CONTROL', 'PITMIN')))
    controller_edit(xml_file_path, 'Channel', 'P_PositiveFeatherAngle',
                    str(get_block(in_file_path, 'CONTROL', 'PITMAX')))
    controller_edit(xml_file_path, 'Channel', 'P_MinimumPitchRateLimit',
                    str(get_block(in_file_path, 'CONTROL', 'PITRMIN')))
    controller_edit(xml_file_path, 'Channel', 'P_MaximumPitchRateLimit',
                    str(get_block(in_file_path, 'CONTROL', 'PITRMAX')))

    controller_edit(xml_file_path, 'Channel', 'P_YawErrorDeratingPower',
                    str(0.68 * rated_shaft_power))

    controller_edit(xml_file_path, 'Channel', 'P_LowSpeedSyncSpeed',
                    str(min_speed))
    controller_edit(xml_file_path, 'Channel', 'P_LowWindSpeedGenSpeed',
                    str(min_speed))

    controller_edit(xml_file_path, 'Channel', 'P_MaxGeneratorSpeedForSwTrip',
                    str(1.18 * rated_speed))
    controller_edit(xml_file_path, 'Channel', 'P_PitchSanityGeneratorSpeedLevel',
                    str(1.15 * rated_speed))
    controller_edit(xml_file_path, 'Channel', 'P_CollectivePitchMinRate',
                    str(0.99 * get_block(in_file_path, 'CONTROL', 'PITRMIN')))
    controller_edit(xml_file_path, 'Channel', 'P_LowPitchAngleOverspeedSpeedLevel',
                    str(1.15 * rated_speed))

    controller_edit(xml_file_path, 'Channel', 'P_RotorRadius',
                    str(0.5 * get_block(in_file_path, 'RCON', 'DIAM')))

    controller_edit(xml_file_path, 'Channel', 'P_MaximumGeneratorTorque',
                    str(get_block(in_file_path, 'GENER', 'GTMAX')))

    controller_edit(xml_file_path, 'Channel', 'P_OptimalModeGain',
                    str(get_block(in_file_path, 'CONTROL', 'GAIN_TSR')))

    controller_edit(xml_file_path, 'Channel', 'P_PitchAccelerationLimit',
                    str(get_xml(in_file_path, 'PitchActuator', 'UpperAccelerationLimit')))
    controller_edit(xml_file_path, 'Channel', 'P_UltimateRateLimit',
                    str(get_block(in_file_path, 'CONTROL', 'PITRMAX')))

    controller_edit(xml_file_path, 'Channel', 'P_TowerExcEnabled', 'false')

    genspeed = [str(min_speed),
                str(0.7 * rated_speed),
                str(rated_speed),
                str(1.2 * rated_speed),
                str(1.4 * rated_speed)]

    fstorque = ['0',
                str(rated_torque),
                str(rated_torque),
                str(0.83 * rated_torque),
                '0']

    t_fastshutdowntorque = [genspeed, fstorque]

    controller_edit(xml_file_path, 'Table', 'T_FastShutdownTorque', t_fastshutdowntorque)

    dynamicfinepitch_pos = [str(get_block(in_file_path, 'CONTROL', 'PITMIN')),
                            '0.3491']
    dynamicfinepitch_pit = [str(get_block(in_file_path, 'CONTROL', 'PITMIN')),
                            '0.2095']

    t_dynamicfinepitch = [dynamicfinepitch_pos, dynamicfinepitch_pit]

    controller_edit(xml_file_path, 'Table', 'T_DynamicFinePitch', t_dynamicfinepitch)

    logging(root_dir, ' 控制器模型信息修正完成')
    print('控制器模型信息修正完成')

    return


# subFunction 取得指定目录下的特定类型文件完整路径 # Done
def get_typefile(root_dir, filetype, name = ''):
    file_dirs = []
    for root, dirs, files in os.walk(root_dir, topdown=False):
        for file in files:
            if os.path.splitext(file)[1].lower() == filetype:
                file_dirs.append(os.path.join(root, file))
    if name:
        file_dirs_a = []
        for dir in file_dirs:
            if name in dir:
                file_dirs_a.append(dir)
        file_dirs = file_dirs_a
    return file_dirs


# subFunction 分割提取的字符串抦并取出相应位置数据 # Done
def do_split(line, symbol="", pos=0):
    temp = line.split(symbol)
    return temp[pos]


# subFunction xml格式信息覆盖 # Done
def change_xml(file_dir, parents, child, info):

    parser = etree.XMLParser(strip_cdata=False, ns_clean=True)
    xml_tree = etree.parse(file_dir, parser)
    elem = r'//' + parents + r'//' + child
    xml_tree.xpath(elem)[0].text = info
    xml_tree.write(file_dir, pretty_print=True, encoding='ISO-8859-1', xml_declaration=True)

    return


# subFunction 模块信息覆盖 #  Done
# 针对模块唯一字段替换 #
def change_block(file_dir, start, label, info, end='MEND'):
    iread = 0
    start = 'MSTART ' + start
    with open(file_dir, 'r') as block:
        lines = block.readlines()
        for j in range(len(lines)):
            if start in lines[j]:
                iread = 1
            if iread:
                split_info = lines[j].split('\t')
                if label == split_info[0]:
                    lines[j] = label + '\t' + info + '\n'
                elif end in lines[j]:
                    break
    with open(file_dir, 'w') as block:
        block.writelines(lines)

    return


# subFunction 字段信息覆盖 # Done
# 针对项目唯一字段替换 #
def change_info(file_dir, label, info):

    with open(file_dir, 'r') as target:
        lines = target.readlines()
        for i in range(len(lines)):
            split_info = lines[i].split('\t')
            if label == split_info[0]:
                lines[i] = label + '\t' + info + '\n'
                break
    with open(file_dir, 'w') as target:
        target.writelines(lines)
    return


# subFunction xml格式信息提取 #
def get_xml(file_dir, parents, child):
    parser = etree.XMLParser(strip_cdata=False, ns_clean=True)
    xml_tree = etree.parse(file_dir, parser)
    elem = r'//' + parents + r'//' + child
    return float(xml_tree.xpath(elem)[0].text)


# subFunction 模块信息提取 # Done
# 针对模块唯一字段提取 #
def get_block(file_dir, start, label):
    answer = '0'
    iread = 0
    start = 'MSTART ' + start
    with open(file_dir, 'r') as block:
        lines = block.readlines()
        for j in range(len(lines)):
            if start in lines[j]:
                iread = 1
            if iread:
                split_info = lines[j].split('\t')
                if label == split_info[0]:
                    answer = split_info[1]
                    break
    return float(answer.strip())


# subFunction 字段信息提取 # Done
# 针对项目唯一字段提取 #
def get_info(file_dir, label):
    answer = '0'
    with open(file_dir, 'r') as target:
        lines = target.readlines()
        for i in range(len(lines)):
            split_info = lines[i].split('\t')
            if label == split_info[0]:
                answer = split_info[1]
                break
    return answer.strip()


# subFunction 删除多余xml格式信息 # Done
def delete_xml(file_dir, parents, child, attrib=None, info=None):
    parser = etree.XMLParser(strip_cdata=False, ns_clean=True)
    xml_tree = etree.parse(file_dir, parser)
    elem = r'//' + parents + r'//' + child
    xml_tree.xpath(elem)[0].clear()
    ns = xml_tree.xpath(elem)[0].nsmap

    if attrib:
        addr = ""
        for keys in ns:
            addr = ns[keys]
        xml_tree.xpath(elem)[0].set(etree.QName(addr, attrib), info)

    xml_tree.write(file_dir, pretty_print=True, encoding='ISO-8859-1', xml_declaration=True)
    return


# subFunction 删除多余模块信息 # Done
def delete_block(file_dir, start, end='MEND'):
    iread = 0
    i = 0
    start = 'MSTART ' + start
    with open(file_dir, 'r') as block:
        lines = block.readlines()
        for j in range(len(lines)):
            if start in lines[j]:
                iread = 1
                i = j
            if iread:
                if end in lines[i]:
                    lines.pop(i)
                    break
                else:
                    lines.pop(i)

    with open(file_dir, 'w') as block:
        block.writelines(lines)

    return


# subFunction 删除多余信息 #
def delete_info(file_dir, info):
    with open(file_dir, 'r') as file:
        lines = file.readlines()
        for j in range(len(lines)):
            if info in lines[j]:
                lines.pop(j)
                break

    with open(file_dir, 'w') as block:
        block.writelines(lines)


# subFunction 添加缺失模块信息 # Done
def add_block(file_dir, pos, block):
    pos = 'MSTART ' + pos
    with open(file_dir, 'r') as file:
        lines = file.readlines()
        for j in range(len(lines)):
            if pos in lines[j]:
                for line in reversed(block):
                    lines.insert(j, line)
                break

    with open(file_dir, 'w') as block:
        block.writelines(lines)

    return


# subFunction 添加缺失信息 #
def add_info(file_dir, pos, info):
    with open(file_dir, 'r') as file:
        lines = file.readlines()
        for j in range(len(lines)):
            if pos in lines[j]:
                lines.insert(j, info)
                break

    with open(file_dir, 'w') as block:
        block.writelines(lines)

    return


# subFunction 获得整段模块 #
def catch_block(file_dir, block):
    block_r = ['\n']
    block = 'MSTART ' + block
    iread = 0
    b_end = 'MEND'
    with open(file_dir, 'r') as file:
        lines = file.readlines()
        for j in range(len(lines)):
            if block in lines[j]:
                iread = 1
            if iread:
                block_r.append(lines[j])
                if b_end in lines[j]:
                    break
    return block_r


# subFunction 运行bat文件 # Done
def run_bat(root_dir, sign):
    file_path = os.path.join(root_dir, sign)
    bat = sign + '.bat'
    bat_path = os.path.join(file_path, bat)
    single_run(bat_path)
    return


# subFunction 控制器xml信息替换 # Done
def controller_edit(file_dir, section, label, info):

    parser = etree.XMLParser(strip_cdata=False, ns_clean=True)
    xml_tree = etree.parse(file_dir, parser)

    root_dir = os.path.dirname(os.path.dirname(file_dir))

    if section == 'Channel':
        channel_names = xml_tree.xpath('//Channel/Name')
        for i in range(len(channel_names)):
            if channel_names[i].text == label:
                channel_names[i].xpath('..//InitialValue')[0].text = info

        logging(root_dir, ' SET ' + label + ' TO ' + info)

    if section == 'Table':
        table_names = xml_tree.xpath('//Table/Name')
        for i in range(len(table_names)):
            if table_names[i].text == label:
                for j in range(len(table_names[i].xpath('..//Row'))):
                    for k in range(len(table_names[i].xpath('..//Row')[0].xpath('./Value'))):
                        if j <= (len(info)-1):
                            if k <= (len(info[j])-1):
                                table_names[i].xpath('..//Row')[j].xpath('./Value')[k].text = info[j][k]
                            else:
                                table_names[i].xpath('..//Row')[j].xpath('./Value')[k].text = \
                                    table_names[i].xpath('..//Row')[j].xpath('./Value')[k-1].text
                        else:
                            table_names[i].xpath('..//Row')[j].xpath('./Value')[k].text = '0'
        logging(root_dir, ' SET ' + label + ' BASED ON MODEL INFO')
    xml_tree.write(file_dir, pretty_print=True, encoding='utf-8', xml_declaration=True)

    return


# subFunction 生成日志文件  #
def logging(root_dir, info):
    logfile_path = os.path.join(root_dir, 'log.txt')
    loglines = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())) + info + '\n'
    with open(logfile_path, 'a+') as logit:
        logit.writelines(loglines)


# #######Second part######## #
# subFunction 计算准备 #
def prepare(root_dir):
    config_dir = os.path.join(root_dir, 'config1.txt')
    with open(config_dir, 'r') as file_to_read:
        lines = file_to_read.read()
    dir = lines
    model_code = ['@echo off\n',
                  'cd %~dp0\n',
                  'setlocal EnableDelayedExpansion\n',
                  '''for /f "delims=" %%i in ('"dir /aa/s/b/on *.prj"') do (\n''',
                  'set file=%%~fi\n',
                  'set file=!file:/=/!\n',
                  ')\n',
                  '\n',
                  'set apath=%~dp0\n',
                  '\n',
                  r'"' + dir + r'Bladed_m72.exe" -Prj !file! -RunDir !apath! -ResultPath !apath!']
    cal_code = ['@echo off\n',
                'cd %~dp0\n',
                r'"' + os.path.join(dir, 'dtbladed.exe') + r'"']
    linmod_code = ['@echo off\n',
                   'cd %~dp0\n',
                   r'"' + dir + r'dtlinmod.exe"']

    windfile_dir = mkdir(root_dir, 'WindFile')
    mkdir(root_dir, 'Temp')
    exctrl_dir = mkdir(root_dir, 'Exctrl')
    model_dir = mkbat(root_dir, 'Model', model_code)
    mkbat(root_dir, 'Performance', cal_code)
    mkbat(root_dir, 'Campbell', cal_code)
    mkbat(root_dir, 'Linear', cal_code)
    mkbat(root_dir, 'LinearModel', linmod_code)

    prj_dir = get_typefile(root_dir, '.prj')

    if len(prj_dir) > 1:
        logging(root_dir, ' ERROR : 模型文件过多！')
        tkinter.messagebox.showerror('错误', 'ERROR 201: 模型文件过多！')
        raise Exception('模型文件过多！')

    elif not prj_dir:
        logging(root_dir, ' ERROR : 没有模型文件！')
        tkinter.messagebox.showerror('错误', 'ERROR 202: 没有模型文件！')
        raise Exception('没有模型文件')

    shutil.copy(prj_dir[0], os.path.join(model_dir, 'Model.prj'))

    wind_dir = get_typefile(root_dir, '.xlsx', 'WindFile')

    if len(wind_dir) > 1:
        logging(root_dir, ' ERROR : 风文件过多！')
        tkinter.messagebox.showerror('错误', 'ERROR 203: 风文件过多！')
        raise Exception('风文件过多！')

    elif not wind_dir:
        logging(root_dir, ' ERROR : 没有风文件！')
        tkinter.messagebox.showerror('错误', 'ERROR 204: 没有风文件！')
        raise Exception('没有风文件')

    shutil.copy(wind_dir[0], windfile_dir)

    dll_dirs = get_typefile(root_dir, '.dll', 'Discon')

    xml_dirs = get_typefile(root_dir, '.xml', 'Parameters')

    if len(dll_dirs) > 1:
        logging(root_dir, ' ERROR : 控制器dll文件过多！')
        tkinter.messagebox.showerror('错误', 'ERROR 205: 控制器dll文件过多！')
        raise Exception('控制器dll文件过多！')

    elif not dll_dirs:
        logging(root_dir, ' ERROR : 没有控制器dll文件！')
        tkinter.messagebox.showerror('错误', 'ERROR 206: 没有控制器dll文件！')
        raise Exception('没有控制器dll文件！')

    if len(xml_dirs) > 1:
        logging(root_dir, ' ERROR : 控制器xml文件过多！')
        tkinter.messagebox.showerror('错误', 'ERROR 207: 控制器xml文件过多！')
        raise Exception('控制器xml文件过多！')

    elif not xml_dirs:
        logging(root_dir, ' ERROR : 没有控制器xml文件！')
        tkinter.messagebox.showerror('错误', 'ERROR 208: 没有控制器xml文件！')
        raise Exception('没有控制器xml文件')

    shutil.copy(dll_dirs[0], exctrl_dir)
    shutil.copy(xml_dirs[0], exctrl_dir)


# subFunction 创建响应路径 #
def mkdir(root_dir, folder):
    target = os.path.join(root_dir, folder)
    isexist = os.path.exists(target)
    if isexist:
        shutil.rmtree(target)
    os.makedirs(target)
    return target


# subFunction 创建bat命令 #
def mkbat(root_dir, target, code):
    target_dir = mkdir(root_dir, target)
    target_bat = target + '.bat'
    bat_dir = os.path.join(target_dir, target_bat)
    fp = open(bat_dir, 'w')
    for lines in code:
        fp.writelines(lines)
    fp.close()
    return target_dir


# Function 并行执行全部PID优化计算 #
def pid_cal(root_dir):
    global txt
    genfile = os.path.join(root_dir, 'GenFile.exe')
    single_run(genfile)

    cal_folder = os.path.join(root_dir, 'PIDCal')
    exe_path = get_typefile(cal_folder, '.exe')
    former_range = []
    later_range = []

    for exe in exe_path:
        if 'Pitch1' in exe:
            former_range.append(exe)
        elif 'Torque' in exe:
            former_range.append(exe)
        else:
            later_range.append(exe)

    pool = ThreadPool()
    pool.map(single_run, former_range)
    pool.close()
    pool.join()

    first_pitch = os.path.join(cal_folder, 'Pitch1')
    txt_path = get_typefile(first_pitch, '.txt')
    first_result = ''
    for txt in txt_path:
        if 'Result' in os.path.basename(txt):
            first_result = txt
    with open(first_result, 'r') as first:
        info = first.read()
        first_td = do_split(info, ' ', 3)

    other_station = []
    for exe in exe_path:
        if 'Pitch1' not in exe and 'Torque' not in exe:
            other_dir = os.path.dirname(exe)
            other_txt = get_typefile(other_dir, '.txt')
            for txt in other_txt:
                if 'Station.txt' in txt:
                    other_station.append(txt)

    for txt in other_station:
        with open(txt, 'r') as station:
            info = station.read()
            temp = info.split(' ')
            temp[3] = first_td
            info_f = ' '.join(temp)
        with open(txt, 'w') as station:
            station.write(info_f)

    pool = ThreadPool()
    pool.map(single_run, later_range)
    pool.close()
    pool.join()


# subFunction 执行exe #
def single_run(command, additional=None):
    dir = os.path.dirname(command)
    if not additional:
        final_command = 'cd /d' + dir + '&&' + command
    else:
        final_command = 'cd /d' + dir + '&&' + command + '&&' + additional
    os.system(final_command)


# Function 将Result.txt中的结果汇总写入Station.xlsx #
def get_result(root_dir):
    temp_dir = os.path.join(root_dir, 'Temp')
    station_file = get_typefile(temp_dir, '.xlsx', 'Station')[0]

    station = xlrd.open_workbook(station_file)
    pitch_info = station.sheet_by_name('Pitch')
    torque_info = station.sheet_by_name('Torque')

    station_order = pitch_info.col_values(0)
    pitch_angle = pitch_info.col_values(1)
    wind_speed = pitch_info.col_values(2)

    q_station = torque_info.col_values(0)[0]
    q_pitch_angle = torque_info.col_values(1)[0]
    q_wind_speed = torque_info.col_values(2)[0]

    cal_dir = os.path.join(root_dir, 'PIDCal')
    txt_path = get_typefile(cal_dir, '.txt')
    result_path = []
    for txt in txt_path:
        if 'Result' in os.path.basename(txt) and '_all' not in os.path.basename(txt):
            result_path.append(txt)
    result_path.sort()

    q_pro_gain = 0
    q_int_const = 0
    q_ts = 0
    q_tr = 0

    p_pro_gain = []
    p_int_const = []
    p_der_gain = []
    p_der_const = 0
    p_ts = []
    p_tr = []

    for result in result_path:
        if 'Torque' in result:
            with open(result, 'r') as torque:
                q_result = torque.read()
                temp = q_result.split(' ')
                q_pro_gain = float(temp[0])
                q_int_const = float(temp[1])
                q_ts = float(temp[2])
                q_tr = float(temp[3])
        elif 'Pitch1' in result:
            with open(result, 'r') as pitch:
                p_result = pitch.read()
                temp = p_result.split(' ')
                p_pro_gain.append(float(temp[0]))
                p_int_const.append(float(temp[1]))
                p_der_gain.append(float(temp[2]))
                p_der_const = float(temp[3])
                p_ts.append(float(temp[4]))
                p_tr.append(float(temp[5]))
        elif 'Pitch' in result and 'Pitch1' not in result:
            with open(result, 'r') as pitch:
                p_result = pitch.read()
                temp = p_result.split(' ')
                p_pro_gain.append(float(temp[0]))
                p_int_const.append(float(temp[1]))
                p_der_gain.append(float(temp[2]))
                p_ts.append(float(temp[4]))
                p_tr.append(float(temp[5]))

    station_f = Workbook()
    pitch_sheet = station_f.create_sheet(title = 'Pitch')
    torque_sheet = station_f.create_sheet(title = 'Torque')

    for i in range(len(p_pro_gain)):
        data = []
        data.append(station_order[i])
        data.append(pitch_angle[i])
        data.append(wind_speed[i])
        data.append(p_pro_gain[i])
        data.append(p_int_const[i])
        data.append(p_der_gain[i])
        data.append(p_ts[i])
        data.append(p_tr[i])
        if i == 0:
            data.append(p_der_const)
        pitch_sheet.append(data)

    torque_data = []
    torque_data.append(q_station)
    torque_data.append(q_pitch_angle)
    torque_data.append(q_wind_speed)
    torque_data.append(q_pro_gain)
    torque_data.append(q_int_const)
    torque_data.append(q_ts)
    torque_data.append(q_tr)
    torque_sheet.append(torque_data)

    station_f.save(station_file)


def print_pid_to_xml(root_dir):
    temp_dir = os.path.join(root_dir, 'Temp')
    file_dir = get_typefile(temp_dir, '.xlsx')

    stationfile_dir = ''

    for i in file_dir:
        if 'Station' in os.path.basename(i):
            stationfile_dir = i
            break

    if not stationfile_dir:
        logging(root_dir, ' ERROR : 未找到工作点文件')
        tkinter.messagebox.showerror('错误', 'ERROR 209: 未找到工作点文件！')
        raise Exception('未找到工作点文件')

    station = xlrd.open_workbook(stationfile_dir)
    pitch_info = station.sheet_by_name('Pitch')
    torque_info = station.sheet_by_name('Torque')

    pitch_angle = pitch_info.col_values(1)
    p_pro_gain = pitch_info.col_values(3)
    p_int_const = pitch_info.col_values(4)
    p_der_gain = pitch_info.col_values(5)
    p_der_const = pitch_info.col_values(8)[0]

    q_pro_gain = torque_info.col_values(3)[0]
    q_int_const = torque_info.col_values(4)[0]
    q_int_gain = q_pro_gain/q_int_const

    control_dir = os.path.join(root_dir, 'Exctrl')
    xml_file_path = get_typefile(control_dir, '.xml')[0]

    controller_edit(xml_file_path, 'Channel', 'P_TorqueSpeedProportionalGain',
                            str(q_pro_gain))
    controller_edit(xml_file_path, 'Channel', 'P_TorqueSpeedIntegralGain',
                            str(q_int_gain))

    controller_edit(xml_file_path, 'Channel', 'P_PitchSpeedDerivativeTimeConstant',
                            str(p_der_const))

    t_pitchspeedproportionalgain = [[str(i) for i in pitch_angle if i], [str(i) for i in p_pro_gain if i]]
    t_pitchspeedintegraltimeconstant = [[str(i) for i in pitch_angle if i], [str(i) for i in p_int_const if i]]
    t_pitchspeedderivativegainschedule = [[str(i) for i in pitch_angle if i], [str(i) for i in p_der_gain if i]]

    controller_edit(xml_file_path, 'Table', 'T_PitchSpeedProportionalGain',
                            t_pitchspeedproportionalgain)
    controller_edit(xml_file_path, 'Table', 'T_PitchSpeedIntegralTimeConstant',
                            t_pitchspeedintegraltimeconstant)
    controller_edit(xml_file_path, 'Table', 'T_PitchSpeedDerivativeGainSchedule',
                            t_pitchspeedderivativegainschedule)

    return


def print_filters_to_xml(root_dir):
    temp_dir = os.path.join(root_dir, 'Temp')
    file_dir = get_typefile(temp_dir, '.xlsx')

    filterfile_dir = ''

    for i in range(len(file_dir)):
        if 'Filters' in file_dir[i]:
            filterfile_dir = file_dir[i]
            break

    if not filterfile_dir:
        logging(root_dir, ' ERROR : 未找到滤波器文件')
        tkinter.messagebox.showerror('错误', 'ERROR 210: 未找到滤波器文件！')
        raise Exception('未找到滤波器文件')

    control_dir = os.path.join(root_dir, 'Exctrl')
    xml_file_path = get_typefile(control_dir, '.xml')[0]

    parser = etree.XMLParser(strip_cdata=False, ns_clean=True)
    xml_tree = etree.parse(xml_file_path, parser)

    rated_speed = 1
    channel_names = xml_tree.xpath('//Channel/Name')
    for i in range(len(channel_names)):
        if channel_names[i].text == 'P_RatedGeneratorSpeed':
            rated_speed = channel_names[i].xpath('..//InitialValue')[0].text

    filters = xlrd.open_workbook(filterfile_dir)
    pitch_filters = filters.sheet_by_name('Pitch')
    torque_filters = filters.sheet_by_name('Torque')
    naf_filters = filters.sheet_by_name('NAF')

    f_pitchspeedfilters = filter_set(pitch_filters, 3, rated_speed)
    f_torquespeedfilters = filter_set(torque_filters, 2, rated_speed)
    f_nacelleaccfafilters = filter_set(naf_filters, 3, rated_speed)

    # Handle should be defined by the filters been chose. #
    # Here the filters were semi-auto set then the handle could be set inside the scripts.#

    controller_edit(xml_file_path, 'Table', 'F_PitchSpeedFilters',
                            f_pitchspeedfilters)
    controller_edit(xml_file_path, 'Table', 'F_TorqueSpeedFilters',
                            f_torquespeedfilters)
    controller_edit(xml_file_path, 'Table', 'F_NacelleAccFAFilters',
                            f_nacelleaccfafilters)

    return


# subFunction 转换滤波器格式 #
def filter_set(sheet, handle, rs):
    filters = []

    for i in range(sheet.nrows):
        filters.append([str(i) for i in sheet.row_values(i)])

    for j in range(len(filters)):
        filters[j].insert(0, '1')
        if j == 0:
            filters[j].insert(1, 'ZeroOrder')
        else:
            filters[j].insert(1, 'SecondOrder')
        filters[j].insert(2, 'SecondOrder')
        filters[j].insert(3, '0')
        filters[j].insert(4, '0')
        if len(filters) - j <= handle:
            filters[j].append(str(rs))
        else:
            filters[j].append('0')
        filters[j].append('0')

    return filters


# subFunction Pitch2-end Pareto Front 结果初步筛选 #
def raw_result(file_path):
    Paras = ['Kp', 'Ti', 'Kd', 'Td', 'SettlingTime', 'RiseTime', 'Overshoot', 'Osc']

    df_ori = pandas.read_csv(file_path, header=None, sep='\s+', names=Paras)
    df = df_ori
    df['score'] = 0

    for index, row in df.iterrows():
        if df.loc[index, 'Osc'] == 0:
            point = 0
        elif df.loc[index, 'Osc'] < 3:
            point = 5
        elif df.loc[index, 'Osc'] < 4:
            point = 4
        elif df.loc[index, 'Osc'] < 5:
            point = 3
        elif df.loc[index, 'Osc'] < 6:
            point = 2
        elif df.loc[index, 'Osc'] < 7:
            point = 1
        else:
            point = 0
        df.loc[index, 'score'] = df.loc[index, 'score'] + point

    for index, row in df.iterrows():
        if df.loc[index, 'SettlingTime'] < 15:
            point = 5
        elif df.loc[index, 'SettlingTime'] < 20:
            point = 4
        elif df.loc[index, 'SettlingTime'] < 25:
            point = 3
        elif df.loc[index, 'SettlingTime'] < 30:
            point = 2
        elif df.loc[index, 'SettlingTime'] < 35:
            point = 1
        else:
            point = 0
        df.loc[index, 'score'] = df.loc[index, 'score'] + point

    for index, row in df.iterrows():
        if df.loc[index, 'RiseTime'] < 5:
            point = 5
        elif df.loc[index, 'RiseTime'] < 6:
            point = 4
        elif df.loc[index, 'RiseTime'] < 7:
            point = 3
        elif df.loc[index, 'RiseTime'] < 8:
            point = 2
        elif df.loc[index, 'RiseTime'] < 9:
            point = 1
        else:
            point = 0
        df.loc[index, 'score'] = df.loc[index, 'score'] + point

    for index, row in df.iterrows():
        if df.loc[index, 'Overshoot'] < 20:
            point = 5
        elif df.loc[index, 'Overshoot'] < 25:
            point = 4
        elif df.loc[index, 'Overshoot'] < 30:
            point = 3
        elif df.loc[index, 'Overshoot'] < 35:
            point = 2
        elif df.loc[index, 'Overshoot'] < 40:
            point = 1
        else:
            point = 0
        df.loc[index, 'score'] = df.loc[index, 'score'] + point

    df_r = df.loc[df.score == max(df.score), ['Kp', 'Ti', 'Kd']]

    return df_r


class LoadCal:

    def __init__(self, root, ltype):
        self.root = root
        self.type = ltype
        self.attr = {}
        self.vmap = {}
        self.result = {}
        self.whole_results = {}
        self.target = 0
        self.dir_list = []
        self.type_list = ['DLC12',
                          'DLC13',
                          'DLC14',
                          'DLC15',
                          'DLC21',
                          'DLC22',
                          'DLC23',
                          'DLC24',
                          'DLC41',
                          'DLC42',
                          'DLC51',
                          'DLC61',
                          'DLC62',
                          'DLC63',
                          'DLC64',
                          'DLC71',
                          'DLC81']

        if self.type == 'ALL':
            self.type_list = []
            self.dir = mkdir(self.root, self.type)
        elif isinstance(self.type, str):
            self.type_list.remove(self.type)
            self.dir = mkdir(self.root, self.type)
        elif isinstance(self.type, list):
            path = ''
            for word in self.type:
                self.type_list.remove(word)
                path += word
            self.dir = mkdir(self.root, path)
        else:
            logging(self.root, ' ERROR : 输入工况不合法！')
            tkinter.messagebox.showerror('错误', 'ERROR 301: 输入工况不合法！')
            raise Exception('输入工况不合法！！')

    def add_attr(self, name, value):
        self.attr[name] = value

    def cal_prepare(self):
        loadtable = os.path.join(self.root, 'loadtable.xls')
        shutil.copy(loadtable, self.dir)

        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        load_case = excel.Workbooks.Open(os.path.join(self.dir, 'loadtable.xls'))
        try:
            for sheetname in self.type_list:
                load_case.Worksheets(sheetname).Delete()
            load_case.Save()
        finally:
            load_case.Close(SaveChanges=0)
            del excel

        para_list = []
        name_list = []

        if self.attr:
            for key, value in self.attr.items():
                name_list.append(key)
                para_list.append(value)

            results = [[]]
            for para in para_list:
                results = [x + [y] for x in results for y in para]
        else:
            results = [1]

        infile_dir = os.path.join(os.path.join(self.root, 'Model'), 'DTBLADED.IN')
        prjfile_dir = os.path.join(os.path.join(self.root, 'Model'), 'Model.prj')
        dll_dir = os.path.join(os.path.join(self.root, 'Exctrl'), 'Discon.dll')
        xml_dir = os.path.join(os.path.join(self.root, 'Exctrl'), 'Parameters.xml')
        wind_dir = os.path.join(os.path.join(self.root, 'WindFile'), 'WindFile.xlsx')

        Aspose = os.path.join(self.root, 'Aspose.Cells.dll')

        HPSocket4C_U = os.path.join(self.root, 'HPSocket4C_U.dll')
        HPSocketCS = os.path.join(self.root, 'HPSocketCS.dll')
        LoadAssistant = os.path.join(self.root, 'LoadsAssistant.exe')
        RainflowPostConfig = os.path.join(self.root, 'RainflowPostConfig.txt')
        VrConfig = os.path.join(self.root, 'VrConfig.txt')
        load_input = os.path.join(self.root, 'load_input.xlsx')

        loadtable = os.path.join(self.dir, 'loadtable.xls')

        PrdDB = os.path.join(self.root, 'PrdDB.db')
        TowerDesign = os.path.join(self.root, 'TowerDesignCMD.exe')

        for i in range(len(results)):
            cur = 'RUN_' + str(i).zfill(3)
            cur_dir = mkdir(self.dir, cur)

            temp_dict = {}
            for j in range(len(name_list)):
                temp_dict[name_list[j]] = results[i][j]
            self.vmap[cur] = temp_dict

            shutil.copy(infile_dir, cur_dir)
            shutil.copy(prjfile_dir, os.path.join(cur_dir, 'pjmodel.prj'))
            shutil.copy(dll_dir, cur_dir)
            shutil.copy(xml_dir, cur_dir)
            shutil.copy(wind_dir, os.path.join(cur_dir, 'winddata.xlsx'))
            shutil.copy(Aspose, cur_dir)
            shutil.copy(HPSocket4C_U, cur_dir)
            shutil.copy(HPSocketCS, cur_dir)
            shutil.copy(LoadAssistant, cur_dir)
            shutil.copy(RainflowPostConfig, cur_dir)
            shutil.copy(VrConfig, cur_dir)
            shutil.copy(load_input, cur_dir)
            shutil.copy(loadtable, cur_dir)
            shutil.copy(PrdDB, cur_dir)
            shutil.copy(TowerDesign, cur_dir)

            if self.attr:
                for j in range(len(name_list)):
                    name = name_list[j]
                    if 'P_' in name:
                        controller_edit(os.path.join(cur_dir, 'Parameters.xml'), 'Channel', name, results[i][j])
                    elif 'T_' in name:
                        controller_edit(os.path.join(cur_dir, 'Parameters.xml'), 'Table', name, results[i][j])

    def do_cal(self):

        load_exe = get_typefile(self.dir, '.exe', 'LoadsAssistant')

        pool = ThreadPool()
        pool.map(single_run, load_exe)
        pool.close()
        pool.join()

        demo_bat = get_typefile(self.dir, '.bat', 'demo')
        for demo in demo_bat:
            mkdir(os.path.dirname(demo), 'tower')
            single_run(demo)

    # 仅用于优化工况输出塔架与对应参数 #
    def opt_result(self):
        tower_info = get_typefile(self.dir, '.csv', 'TowerInfos')
        tower_weight = []
        run_name = []

        self.whole_results = self.vmap

        for weight_file in tower_info:
            with open(weight_file, 'r') as weight_info:
                info = csv.reader(weight_info)
                weight = 9999
                for row in info:
                    if row[0] == '总重':
                        weight = float(row[1])
                try:
                    if weight == 9999:
                        err_msg = weight_file + '中找不到塔架重量'
                        tkinter.messagebox.showwarning('警告', err_msg)
                        logging(self.root, err_msg)
                        raise Exception(err_msg)
                finally:
                    tower_weight.append(weight)

                run_name.append(os.path.basename(os.path.dirname(weight_file).strip(r'\tower\Result')))
                self.whole_results[run_name]['Weight'] = weight

        best_weight = min(tower_weight)
        best_run = run_name[tower_weight.index(best_weight)]
        best_para = self.vmap[best_run]
        self.result = {'Weight': best_weight,
                       'Run': best_run,
                       'Para': best_para}

    # 输出优化结果至文件 #
    # 依据最优结果修正控制器 #
    def opt_result_collect(self):
        opt_result = self.result
        all_result = self.whole_results

        best_log_file = os.path.join(self.dir, 'best.txt')
        opt_log_file = os.path.join(self.dir, 'opt_log.txt')

        best_log = ['Weight:' + opt_result['Weight'],
                    'Run:' + opt_result['Run']]

        exctrl_dir = os.path.join(os.path.join(self.root, 'Exctrl'), 'Parameters.xml')
        for key, value in opt_result['Para']:
            best_log.append(key + ':' + value)
            if 'P_' in key:
                controller_edit(exctrl_dir, 'Channel', key, value)
            elif 'T_' in key:
                controller_edit(exctrl_dir, 'Table', key, value)
            else:
                pass

        with open(best_log_file, 'w') as file:
            file.writelines(best_log)

        all = pandas.DataFrame(all_result)
        all_in_list = all.values.tolist()

        with open(opt_log_file, 'w') as opt_file:
            opt_file.writelines(all.columns.tolist())
            for s_list in all_in_list:
                opt_file.writelines(s_list)


# Function 基于载荷/塔架重量的控制参数优化 #
# 分工况计算：DLC14/DLC23/DLC42 #
# 过滤失败计算 #
# 返回结果包括最优组及全部计算结果 #
# 依据最优结果修正控制器 #
def load_cal_dlc14(root_dir):
    dlc14 = LoadCal(root_dir, 'DLC14')

    dlc14_name_1 = 'P_NacelleAccFAGain'
    dlc14_value_1 = ['0.03', '0.0375', '0.045', '0.0525', '0.06']
    dlc14.add_attr(dlc14_name_1, dlc14_value_1)

    dlc14_name_2 = 'P_DynamicFinePitchTimeConstant'
    dlc14_value_2 = ['3', '5', '15']
    dlc14.add_attr(dlc14_name_2, dlc14_value_2)

    dlc14_name_3 = 'P_NacelleAccMaxYawError'
    dlc14_value_3 = ['0.349', '0.698', '1.047']
    dlc14.add_attr(dlc14_name_3, dlc14_value_3)

    dlc14_name_4 = 'T_DynamicFinePitch'
    in_dir = os.path.join(root_dir, 'Model')
    in_file_path = get_typefile(in_dir, '.in')[0]
    min_pitch = str(get_block(in_file_path, 'CONTROL', 'PITRMIN'))
    row_1 = [min_pitch, '0.3491']
    dlc14_value_4 = [[row_1, [min_pitch, '0.1396']],
                     [row_1, [min_pitch, '0.1745']],
                     [row_1, [min_pitch, '0.2094']],
                     [row_1, [min_pitch, '0.2443']],
                     [row_1, [min_pitch, '0.0.2793']]]
    dlc14.add_attr(dlc14_name_4, dlc14_value_4)

    dlc14.cal_prepare()
    dlc14.do_cal()
    dlc14.opt_result()
    dlc14.opt_result_collect()


def load_cal_dlc42(root_dir):
    dlc42 = LoadCal(root_dir, 'DLC42')

    dlc42_name_1 = 'P_NormalStopMinPitchRate'
    dlc42_value_1 = ['0.0087', '0.01745', '0.02618', '0.03491']
    dlc42.add_attr(dlc42_name_1, dlc42_value_1)

    dlc42_name_2 = 'P_ShaftPowerTargetRateLimit'
    dlc42_value_2 = ['25000', '35000', '55000', '75000', '90000', '120000']
    dlc42.add_attr(dlc42_name_2, dlc42_value_2)

    dlc42.cal_prepare()
    dlc42.do_cal()
    dlc42.opt_result()
    dlc42.opt_result_collect()


def load_cal_dlc23(root_dir):
    dlc23 = LoadCal(root_dir, 'DLC23')
    dlc23_name_1 = 'T_GridLossShutdownPitchRate'

    gridlossrate_1 = ['0.03491', '0.04363', '0.05236', '0.06109', '0.06981']
    gridlossrate_2 = ['0.00873', '0.01745', '0.02618', '0.03491']

    grid_timerange = [['0', '0.6', '0.62', '5', '5.02', '15'],
                  ['0', '1', '1.02', '5', '5.02', '15'],
                  ['0', '1.4', '1.42', '5', '5.02', '15']]
    grid_raterange = [[rate1, rate1, rate2, rate2, '0.0349', '0.0349']
                      for rate1 in gridlossrate_1 for rate2 in gridlossrate_2]
    dlc23_value_1 = [[row1, row2] for row1 in grid_timerange for row2 in grid_raterange]
    dlc23.add_attr(dlc23_name_1, dlc23_value_1)

    dlc23.cal_prepare()
    dlc23.do_cal()
    dlc23.opt_result()
    dlc23.opt_result_collect()


# 优化工况总成计算 #
# 计算后应依据最终载荷更换塔架 #
def load_cal_ex(root_dir):
    load_cases = ['DLC14', 'DLC42', 'DLC23']
    ex_load = LoadCal(root_dir, load_cases)

    ex_load.cal_prepare()
    ex_load.do_cal()

    tower_process_dir = mkdir(root_dir, 'TowerReplace')
    run_dir = os.path.join(ex_load.dir, 'RUN_000')

    UpdateTowerModel = os.path.join(root_dir, 'UpdateTowerModel.exe')
    prj_dir = os.path.join(run_dir, 'pjmodel.prj')
    in_dir = os.path.join(run_dir, 'DTBLADED.IN')
    csv_dir = os.path.join(run_dir, r'tower\Result\File4Loads.csv')

    shutil.copy(UpdateTowerModel, tower_process_dir)
    shutil.copy(prj_dir, tower_process_dir)
    shutil.copy(in_dir, tower_process_dir)
    shutil.copy(csv_dir, tower_process_dir)

    single_run(os.path.join(tower_process_dir, 'UpdateTowerModel.exe'))

    shutil.copy(os.path.join(tower_process_dir, 'pjmodel.prj'), os.path.join(root_dir, r'Model\Model.prj'))
    shutil.copy(os.path.join(tower_process_dir, 'DTBLADED.IN'), os.path.join(root_dir, r'Model\DTBLADED.IN'))


# 全工况计算使用更新塔架后的模型 #
# 计算后输出最终重量并进行校核 #
def load_cal_all(root_dir):
    load_all = LoadCal(root_dir, 'ALL')

    load_all.cal_prepare()
    load_all.do_cal()

    weight_file = get_typefile(load_all.dir, '.csv', 'TowerInfos')[0]
    tower_weight = 0

    with open(weight_file, 'r') as weight_info:
        info = csv.reader(weight_info)
        weight = 9999
        for row in info:
            if row[0] == '总重':
                weight = float(row[1])
        try:
            if weight == 9999:
                err_msg = weight_file + '中找不到塔架重量'
                tkinter.messagebox.showwarning('警告', err_msg)
                logging(load_all.root, err_msg)
                raise Exception(err_msg)
        finally:
            tower_weight = weight

    comopent_check = mkdir(root_dir, 'VerificationComopent')
    mechanical_check = mkdir(root_dir, 'VerificationMechanical')

    ToPL = os.path.join(root_dir, 'ToPL.exe')
    Data_2X = os.path.join(root_dir, 'Data_2X.pkl')
    load_compare = os.path.join(root_dir, 'load_compare.exe')
    input_structure = os.path.join(root_dir, 'input_structure.xlsx')

    load_input = os.path.join(load_all.dir, r'RUN_000\load_input.xlsx')

    shutil.copy(ToPL, comopent_check)
    shutil.copy(Data_2X, comopent_check)
    shutil.copy(load_input, comopent_check)

    single_run(os.path.join(comopent_check, 'ToPL.exe'))

    shutil.copy(load_compare, mechanical_check)
    shutil.copy(input_structure, mechanical_check)
    shutil.copy(load_input, mechanical_check)

    single_run(os.path.join(mechanical_check, 'load_compare.exe'))

    component_result_dir = os.path.join(comopent_check, 'Check_Cast_Bolt.txt')
    mechanical_result_dir = os.path.join(mechanical_check, 'Check Result Summary.txt')

    if not os.path.exists(component_result_dir):
        logging(root_dir, ' ERROR : 没有强度校核结果文件！')
        tkinter.messagebox.showerror('错误', 'ERROR 401: 没有强度校核结果文件！')
        try:
            raise Exception('没有强度校核结果文件')
        finally:
            pass

    if not os.path.exists(mechanical_result_dir):
        logging(root_dir, ' ERROR : 没有机械校核结果文件！')
        tkinter.messagebox.showerror('错误', 'ERROR 401: 没有机械校核结果文件！')
        try:
            raise Exception('没有机械校核结果文件')
        finally:
            pass

    with open(component_result_dir, 'r') as component:
        c_lines = component.read()
    component_result = c_lines

    with open(mechanical_result_dir, 'r') as mechanical:
        m_lines = mechanical.readlines()

    mechanical_result = m_lines[1]

    result = [str(tower_weight) + '\n', component_result, mechanical_result]

    final_result_dir = os.path.join(root_dir, r'Result\result.txt')

    if os.path.exists(final_result_dir):
        shutil.rmtree(final_result_dir)
        mkdir(root_dir, 'Result')

    with open(final_result_dir, 'w') as final:
        final.writelines(result)


